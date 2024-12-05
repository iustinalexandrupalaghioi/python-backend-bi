from xml.etree.ElementTree import tostring
from flask import Flask, send_file, request, jsonify
from flask_cors import CORS
import asyncpg
import os
from dotenv import load_dotenv
import logging
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.drawing.text import  ParagraphProperties, CharacterProperties
from io import BytesIO
import numpy as np
from scipy.optimize import curve_fit


# Set up logging
logging.basicConfig(level=logging.DEBUG)

# Load environment variables
load_dotenv()

app = Flask(__name__)
CORS(app)

# Database connection URL
DB_URL = os.getenv("DB_URL")



#T1. Fetch sales trend and estimate sales trend when applying for discounts
@app.get("/api/sales/fetch-sales-trend")
async def fetch_sales_with_discounts():
    try:
        logging.debug("Establishing database connection...")
        connection = await asyncpg.connect(dsn=DB_URL)

        # Validate and parse query parameters
        start_date = request.args.get("startDate")
        end_date = request.args.get("endDate")
        gender = request.args.get("gender", "All")
        min_age = request.args.get("ageMin")
        max_age = request.args.get("ageMax")
        city = request.args.get("city", "All")
        trend_type = request.args.get("trendType", "linear")
        frequency = request.args.get("frequency", "Daily").capitalize()  # Normalize capitalization
        prediction_points = int(request.args.get("predictionPoints", 0))

        # Ensure required parameters are present
        if not start_date or not end_date:
            return {"error": "startDate and endDate are required."}, 400

        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            if start_date > end_date:
                return {"error": "startDate must be before endDate."}, 400
        except ValueError:
            return {"error": "Invalid date format. Use YYYY-MM-DD."}, 400

        # Handle min_age and max_age with default to None
        min_age = int(min_age) if min_age and min_age.isdigit() else None
        max_age = int(max_age) if max_age and max_age.isdigit() else None

        # Validate frequency
        valid_frequencies = {"Daily": "day", "Monthly": "month", "Yearly": "year"}
        if frequency not in valid_frequencies:
            return {"error": f"Invalid frequency. Choose from {list(valid_frequencies.keys())}."}, 400

        # Build the query to fetch sales data with discounts
        date_trunc_unit = valid_frequencies[frequency]
        query = f"""
            SELECT 
                DATE_TRUNC('{date_trunc_unit}', sale_date) AS period,
                discount_name,
                discount_rate,
                SUM(total_price) AS total_sales
            FROM Sales
            LEFT JOIN Discounts ON Sales.discount_id = Discounts.discount_id
            LEFT JOIN Clients ON Sales.client_id = Clients.client_id
            LEFT JOIN Cities ON Sales.city_id = Cities.city_id
            WHERE sale_date BETWEEN $1 AND $2 AND sales.discount_id IS NOT NULL
        """
        params = [start_date, end_date]

        if gender != "All":
            query += " AND gender = $3"
            params.append(gender)
        if min_age is not None:
            query += f" AND age >= ${len(params) + 1}"
            params.append(min_age)
        if max_age is not None:
            query += f" AND age <= ${len(params) + 1}"
            params.append(max_age)
        if city != "All":
            query += f" AND city_name = ${len(params) + 1}"
            params.append(city)

        query += " GROUP BY period, discount_name, discount_rate ORDER BY period, discount_name;"

        result = await connection.fetch(query, *params)

        if not result:
            return {"error": "No sales data found with discounts for the specified range."}, 404

        # Prepare sales data, grouped by discount
        sales_data = {}
        for row in result:
            discount = float(row["discount_rate"])
            friendly_name = row["discount_name"] + ": " + str(discount)
            sale_date = row["period"].strftime("%Y-%m-%d")
            total_sales = float(row["total_sales"])
            if discount not in sales_data:
                sales_data[discount] = {"dates": [], "sales": []}
            sales_data[discount]["dates"].append(sale_date)
            sales_data[discount]["sales"].append(total_sales)

        # Prepare trend data (optional)
        trend_data = {}
        for friendly_name, data in sales_data.items():
            trend_line = []
            future_trend = []

            # Ensure trend_type and prediction_points are set correctly
            if trend_type and prediction_points > 0:
                x_data = np.arange(len(data["dates"]))  # X axis: [0, 1, 2, ..., len(data["dates"])-1]
                y_data = np.array(data["sales"])        # Y axis: sales data

                # Calculate trend line and future trend
                trend_line, future_trend = calculate_trend(x_data, y_data, trend_type, prediction_points)

            # Prepare the trend data, ensuring no out-of-bounds access occurs
            trend_data[friendly_name] = {
                "trend": [{"date": data["dates"][i], "trend_value": trend_line[i]} for i in range(len(trend_line))] if trend_line.any() else [],
                "future_trend": [
                    {"date": (end_date + timedelta(days=i)).strftime("%Y-%m-%d"), "trend_value": future_trend[i]}
                    for i in range(len(future_trend))  # Avoid accessing out-of-bounds
                ] if future_trend.any() else []
            }

        return {"trend_data": trend_data}

    except Exception as e:
        logging.error(f"Error: {e}")
        return {"error": f"An error occurred while processing the request: {e}"}, 500

    finally:
        await connection.close()


        
        

#2. Exponential function for curve fitting
def exponential_func(x, a, b, c):
    """Exponential function: a * exp(b * x) + c."""
    return a * np.exp(b * x) + c

#2. Trend calculation for different types
def calculate_trend(x_data, y_data, trend_type, prediction_points):
    """Calculate trend line and future predictions based on the trend type."""
    future_x_data = np.arange(len(x_data) + prediction_points)
    
    if trend_type == "linear":
        coeffs = np.polyfit(x_data, y_data, 1)
        trend_line = np.polyval(coeffs, x_data)
        future_trend = np.polyval(coeffs, future_x_data)
    elif trend_type == "exponential":
        params, _ = curve_fit(exponential_func, x_data, y_data, p0=(1, 0.01, 1), maxfev=2000)
        trend_line = exponential_func(x_data, *params)
        future_trend = exponential_func(future_x_data, *params)
    elif trend_type == "polynomial":
        coeffs = np.polyfit(x_data, y_data, 2)
        trend_line = np.polyval(coeffs, x_data)
        future_trend = np.polyval(coeffs, future_x_data)
    elif trend_type == "logarithmic":
        coeffs, _ = curve_fit(lambda x, a, b: a * np.log(x + 1) + b, x_data, y_data, p0=(1, 1))
        trend_line = coeffs[0] * np.log(x_data + 1) + coeffs[1]
        future_trend = coeffs[0] * np.log(future_x_data + 1) + coeffs[1]
    elif trend_type == "power-law":
        coeffs, _ = curve_fit(lambda x, a, b: a * x**b, x_data + 1, y_data, p0=(1, 1))
        trend_line = coeffs[0] * (x_data + 1) ** coeffs[1]
        future_trend = coeffs[0] * (future_x_data + 1) ** coeffs[1]
    elif trend_type == "moving_average":
        window_size = 3
        trend_line = np.convolve(y_data, np.ones(window_size) / window_size, mode="same")
        future_trend = np.concatenate([trend_line, np.repeat(trend_line[-1], prediction_points)])
    else:
        raise ValueError(f"Invalid trendType: {trend_type}")
    
    return trend_line, future_trend


#2. export highs and lows of sales + trend
def create_excel_report(dates, sales, trend_line, future_trend, frequency, prediction_points, end_date):
    """Create an Excel workbook with an enhanced sales trend chart."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales Data"

    # Add headers
    sheet.append(["Date", "Total Sales", "Trend Line", "Estimated Future Trend"])

    # Add actual data rows
    for i, (date, sale) in enumerate(zip(dates, sales)):
        sheet.append([date, sale, trend_line[i], None])

    # Add future prediction rows
    future_dates = [
        end_date + timedelta(days=i) if frequency == "Daily" else
        end_date + timedelta(days=30 * i) if frequency == "Monthly" else
        end_date + timedelta(days=365 * i)
        for i in range(1, prediction_points + 1)
    ]
    for i, future_date in enumerate(future_dates):
        sheet.append([future_date, None, None, future_trend[len(dates) + i]])

    # Create and add a line chart
    chart = LineChart()
    chart.title = "Sales trend"
    chart.y_axis.title = "Total sales"
    chart.x_axis.title = "Date"
    chart.style = 10
    chart.x_axis.number_format = 'yyyy-mm-dd'
    chart.x_axis.title = "Date"
    
    # Define data and labels for the chart
    data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row, max_col=4)
    labels = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)

    # Add data to the chart and set categories
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    # Increase chart size
    chart.width = 25
    chart.height = 12

    # Add the chart to the sheet
    sheet.add_chart(chart, "G3")

    # Save workbook to a BytesIO stream
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
   
   
#2. API endpoint to fetch sales
@app.get("/api/sales/fetch-sales")
async def fetch_sales():
    try:
        logging.debug("Establishing database connection...")
        connection = await asyncpg.connect(dsn=DB_URL)

        # Validate and parse query parameters
        start_date = request.args.get("startDate")
        end_date = request.args.get("endDate")
        gender = request.args.get("gender", "All")
        min_age = request.args.get("minAge")
        max_age = request.args.get("maxAge")
        city = request.args.get("city", "All")

        # Ensure required parameters are present
        if not start_date or not end_date:
            return {"error": "startDate and endDate are required."}, 400

        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            if start_date > end_date:
                return {"error": "startDate must be before endDate."}, 400
        except ValueError:
            return {"error": "Invalid date format. Use YYYY-MM-DD."}, 400

        min_age = int(min_age) if min_age else None
        max_age = int(max_age) if max_age else None

        # Build and execute query
        query = """
            SELECT sale_id, title, age_group_name, description, age, gender, sale_date, quantity, total_price AS total_sales, category_name, city_name
            FROM Sales
            LEFT JOIN Clients ON Sales.client_id = Clients.client_id
            LEFT JOIN Age_groups ON Clients.age_group_id = Age_groups.age_group_id
            LEFT JOIN Books ON Sales.book_id = Books.book_id
            LEFT JOIN Subcategories ON Books.subcategory_id = Subcategories.subcategory_id
            LEFT JOIN Categories ON Subcategories.category_id = Categories.category_id
            LEFT JOIN Cities ON Sales.city_id = Cities.city_id
            WHERE sale_date BETWEEN $1 AND $2
        """
        params = [start_date, end_date]

        if gender != "All":
            query += " AND gender = $3"
            params.append(gender)
        if min_age is not None:
            query += f" AND age >= ${len(params) + 1}"
            params.append(min_age)
        if max_age is not None:
            query += f" AND age <= ${len(params) + 1}"
            params.append(max_age)
        if city != "All":
            query += " AND city_name = $3"
            params.append(city)

        query += " ORDER BY sale_date;"
        result = await connection.fetch(query, *params)

        if not result:
            return {"error": "No sales data found for the specified range."}, 404

        # Prepare the response data with additional details
        sales_data = [
            {
                "sale_id": row["sale_id"],
                "book_title": row["title"],
                "age_group": row["age_group_name"],
                "age_group_description": row["description"],
                "age": row["age"],
                "gender": row["gender"],
                "sale_date": row["sale_date"].strftime("%Y-%m-%d"),
                "quantity": row["quantity"],
                "total_sales": row["total_sales"],
                "category": row["category_name"],
                "city": row["city_name"],
            }
            for row in result
        ]

        # Return sales data (for the fetch-sales endpoint)
        return {"data": sales_data}

    except Exception as e:
        logging.error(f"Error: {e}")
        return {"error": f"An error occurred while processing the request: {e}"}, 500

    finally:
        await connection.close()


#2. API endpoint to export sales data and generate report
@app.get("/api/sales/export-sales")
async def export_sales():
    try:
        logging.debug("Establishing database connection...")
        connection = await asyncpg.connect(dsn=DB_URL)

        # Validate and parse query parameters
        start_date = request.args.get("startDate")
        end_date = request.args.get("endDate")
        gender = request.args.get("gender", "All")
        min_age = request.args.get("minAge")
        max_age = request.args.get("maxAge")
        city = request.args.get("city", "All")
        trend_type = request.args.get("trendType", "linear")
        frequency = request.args.get("frequency", "Daily").capitalize()  # Normalize capitalization

        # Ensure required parameters are present
        if not start_date or not end_date:
            return {"error": "startDate and endDate are required."}, 400

        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            if start_date > end_date:
                return {"error": "startDate must be before endDate."}, 400
        except ValueError:
            return {"error": "Invalid date format. Use YYYY-MM-DD."}, 400

        min_age = int(min_age) if min_age else None
        max_age = int(max_age) if max_age else None

        # Validate frequency
        valid_frequencies = {"Daily": "day", "Monthly": "month", "Yearly": "year"}
        if frequency not in valid_frequencies:
            return {"error": f"Invalid frequency. Choose from {list(valid_frequencies.keys())}."}, 400

        # Build and execute query
        date_trunc_unit = valid_frequencies[frequency]
        query = f"""
            SELECT 
                DATE_TRUNC('{date_trunc_unit}', sale_date) AS period,
                SUM(total_price) AS total_sales
            FROM Sales
            LEFT JOIN Clients ON Sales.client_id = Clients.client_id
            LEFT JOIN Age_groups ON Clients.age_group_id = Age_groups.age_group_id
            LEFT JOIN Cities ON Sales.city_id = Cities.city_id
            WHERE sale_date BETWEEN $1 AND $2
        """
        params = [start_date, end_date]

        if gender != "All":
            query += " AND gender = $3"
            params.append(gender)
        if min_age is not None:
            query += f" AND age >= ${len(params) + 1}"
            params.append(min_age)
        if max_age is not None:
            query += f" AND age <= ${len(params) + 1}"
            params.append(max_age)
        if city != "All":
            query += f" AND city_name = ${len(params) + 1}"
            params.append(city)

        query += " GROUP BY period ORDER BY period;"
        result = await connection.fetch(query, *params)

        if not result:
            return {"error": "No sales data found for the specified range."}, 404

        # Prepare the aggregated sales data
        sales_data = [
            {
                "period": row["period"].strftime("%Y-%m-%d"),
                "total_sales": float(row["total_sales"]),
            }
            for row in result
        ]

        # Generate trend line and future predictions
        periods = [row["period"] for row in sales_data]
        sales = [row["total_sales"] for row in sales_data]
        trend_line, future_trend = calculate_trend(np.arange(len(sales)), sales, trend_type, len(sales))

        # Create Excel file with trend chart
        excel_output = create_excel_report(periods, sales, trend_line, future_trend, frequency, len(sales), end_date)

        # Send Excel file as response
        return send_file(
            excel_output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"sales_trend_{frequency}.xlsx"
        )

    except Exception as e:
        logging.error(f"Error: {e}")
        return {"error": f"An error occurred while processing the request: {e}"}, 500

    finally:
        await connection.close()




#1. API endpoint to fetch all categories
@app.get("/api/sales/categories")
async def fetch_categories():
    try:
        # Establish database connection
        conn = await asyncpg.connect(DB_URL)
        try:
            # Fetch all categories
            query = "SELECT category_id, category_name FROM categories ORDER BY category_name;"
            rows = await conn.fetch(query)
            
            # Convert rows to a list of dictionaries
            categories = [dict(row) for row in rows]
            return jsonify(categories)
        finally:
            await conn.close()
    except Exception as e:
        return jsonify({"error": str(e)}), 500

#1. API endpoint to fetch sales per subcategory filtering by category
@app.get("/api/sales/subcategory-series")
async def get_sales_per_subcategory():
    # Get query parameters
    gender = request.args.get("gender", None)
    age_min = request.args.get("ageMin", None, type=int)
    age_max = request.args.get("ageMax", None, type=int)
    start_date = request.args.get("startDate", None)
    end_date = request.args.get("endDate", None)
    category = request.args.get("category", None, type=int)

    # Validate and parse date inputs
    if not start_date or not end_date:
        return jsonify({"error": "startDate and endDate are required."}), 400

    try:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"error": "Invalid date format. Use YYYY-MM-DD."}), 400

    # Base SQL query
    base_query = """
        SELECT 
            sub.subcategory_name AS subcategory_name,
            COUNT(s.sale_id) AS total_sales
        FROM 
            sales s
        INNER JOIN 
            books b ON s.book_id = b.book_id
        INNER JOIN 
            subcategories sub ON b.subcategory_id = sub.subcategory_id
        INNER JOIN 
            categories cat ON sub.category_id = cat.category_id
        INNER JOIN 
            clients c ON s.client_id = c.client_id
        WHERE 
            s.sale_date BETWEEN $1 AND $2
    """
    params = [start_date_obj, end_date_obj]
    conditions = []

    # Add gender filter if not "All"
    if gender and gender.lower() != "all":
        conditions.append(f"c.gender = ${len(params) + 1}")
        params.append(gender)

    # Add age filters if provided
    if age_min is not None:
        conditions.append(f"c.age >= ${len(params) + 1}")
        params.append(age_min)
    if age_max is not None:
        conditions.append(f"c.age <= ${len(params) + 1}")
        params.append(age_max)

    # Add category filter if not "All"
    if category and category != 0:
        conditions.append(f"cat.category_id = ${len(params) + 1}")
        params.append(category)

    # Append conditions to query
    if conditions:
        base_query += " AND " + " AND ".join(conditions)

    # Add grouping and sorting
    base_query += " GROUP BY sub.subcategory_name ORDER BY sub.subcategory_name;"

    # Execute the query
    try:
        conn = await asyncpg.connect(DB_URL)
        try:
            rows = await conn.fetch(base_query, *params)
            data = [dict(row) for row in rows]
            logging.debug(f"Executing query 1: {base_query} with params: {params}")
            return jsonify(data)
        finally:
            await conn.close()
    except Exception as e:
        logging.error(f"Error fetching data: {e}")
        return jsonify({"error": str(e)}), 500

#1. Export bar chart per subcategory filtering by categories
@app.get("/api/sales/export-subcategory-bar-chart")
async def export_sales_per_subcategory_with_bar_chart():
      # Get query parameters
    gender = request.args.get("gender", None)
    age_min = request.args.get("ageMin", None, type=int)
    age_max = request.args.get("ageMax", None, type=int)
    start_date = request.args.get("startDate", None)
    end_date = request.args.get("endDate", None)
    category = request.args.get("category", None, type=int)

    # Validate and parse date inputs
    if not start_date or not end_date:
        return jsonify({"error": "startDate and endDate are required."}), 400

    try:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"error": "Invalid date format. Use YYYY-MM-DD."}), 400

    # Base SQL query
    base_query = """
        SELECT 
            sub.subcategory_name AS subcategory_name,
            COUNT(s.sale_id) AS total_sales
        FROM 
            sales s
        INNER JOIN 
            books b ON s.book_id = b.book_id
        INNER JOIN 
            subcategories sub ON b.subcategory_id = sub.subcategory_id
        INNER JOIN 
            categories cat ON sub.category_id = cat.category_id
        INNER JOIN 
            clients c ON s.client_id = c.client_id
        WHERE 
            s.sale_date BETWEEN $1 AND $2
    """
    params = [start_date_obj, end_date_obj]
    conditions = []

    # Add gender filter if not "All"
    if gender and gender.lower() != "all":
        conditions.append(f"c.gender = ${len(params) + 1}")
        params.append(gender)

    # Add age filters if provided
    if age_min is not None:
        conditions.append(f"c.age >= ${len(params) + 1}")
        params.append(age_min)
    if age_max is not None:
        conditions.append(f"c.age <= ${len(params) + 1}")
        params.append(age_max)

    # Add category filter if not "All"
    if category and category != 0:
        conditions.append(f"cat.category_id = ${len(params) + 1}")
        params.append(category)

    # Append conditions to query
    if conditions:
        base_query += " AND " + " AND ".join(conditions)

    # Add grouping and sorting
    base_query += " GROUP BY sub.subcategory_name ORDER BY sub.subcategory_name;"

    # Execute the query
    try:
        conn = await asyncpg.connect(DB_URL)
        try:
            rows = await conn.fetch(base_query, *params)
            data = [dict(row) for row in rows]
            logging.debug(f"Executing query: {base_query} with params: {params}")

            # Prepare data for the Excel file
            subcategories = [row["subcategory_name"] for row in data]
            total_sales = [row["total_sales"] for row in data]
            
            if not subcategories or not total_sales:
                logging.error("Missing data for subcategories or total_sales.")
                return jsonify({"error": "Missing data for subcategories or sales."}), 400

            # Call the function to generate the Excel file with a bar chart
            excel_output = create_excel_with_bar_chart(subcategories, total_sales)

            # Return the Excel file as a response for download
            return send_file(
                excel_output,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="sales_per_subcategory.xlsx"
            )
        finally:
            await conn.close()
    except Exception as e:
        logging.error(f"Error fetching data: {e}")
        return jsonify({"error": str(e)}), 500

#1. Create bar chart in excel
def create_excel_with_bar_chart(subcategories, sales):
    try:
        # Create a new Excel workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sales Data"

        # Add headers to the sheet
        sheet.append(["Subcategory", "Total Sales"])

        # Add actual data rows (Sales)
        for subcategory, sale in zip(subcategories, sales):
            sheet.append([subcategory, sale])

        # Create a Bar chart for the total sales
        chart = BarChart()
        chart.title = "Sales"
        chart.style = 10
        chart.y_axis.title = "Total Sales"
        chart.x_axis.title = "Subcategory"

         # Define the data for the chart (exclude header row for data)
        data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row)

        # Define the categories for the chart (Subcategories)
        categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)

        # Add the data and categories to the chart
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Increase chart size
        chart.width = 25
        chart.height = 12

        # Position the chart on the sheet
        sheet.add_chart(chart, "G5")

        # Save the workbook to a BytesIO stream
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Return the Excel file with the chart
        return output
    except Exception as e:
        logging.error(f"Error creating Excel file with chart: {e}")
        raise e  # Raise the error to be caught in the API controller
    





#3. API endpoint to fetch sales data for event linking, filter by category
@app.get('/api/sales/fetch-event-sales')
async def fetch_event_sales():
    try:
        logging.debug("Establishing database connection...")
        connection = await asyncpg.connect(dsn=DB_URL)

        # Validate and parse query parameters
        start_date = request.args.get("startDate")
        end_date = request.args.get("endDate")
        category = request.args.get("category", None, type=int)

        # Ensure required parameters are present
        if not start_date or not end_date:
            return {"error": "startDate and endDate are required."}, 400

        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            if start_date > end_date:
                return {"error": "startDate must be before endDate."}, 400
        except ValueError:
            return {"error": "Invalid date format. Use YYYY-MM-DD."}, 400

        # Build and execute query to fetch event data and sales
        base_query = """
            SELECT 
                e.event_name,
                cat.category_name as category_name,
                e.start_date, 
                e.end_date, 
                CAST(e.end_date - e.start_date AS INTEGER) + 1 AS duration,
                SUM(s.quantity) AS total_quantity_sold,
                SUM(s.total_price) AS total_sales,
                COUNT(DISTINCT s.book_id) AS unique_books_sold,

                CASE 
                    WHEN (e.end_date - e.start_date) > 0 
                    THEN SUM(s.total_price) / (e.end_date - e.start_date) 
                    ELSE SUM(s.total_price)
                END AS average_sales_per_day,
                
                CASE 
                    WHEN (e.end_date - e.start_date) > 0 
                    THEN SUM(s.quantity) / (e.end_date - e.start_date) 
                    ELSE SUM(s.quantity)
                END AS average_books_sold_per_day
            FROM events e
            LEFT JOIN sales s ON s.event_id = e.event_id
            INNER JOIN books b ON b.book_id = s.book_id
            INNER JOIN subcategories sub ON b.subcategory_id = sub.subcategory_id
            LEFT JOIN categories cat ON sub.category_id = cat.category_id
            WHERE e.start_date BETWEEN $1 AND $2
        """

        # Add category filter dynamically
        params = [start_date, end_date]
        if category and category != 0:
            base_query += " AND cat.category_id = $3"
            params.append(category)

        # Group and order results
        base_query += """
            GROUP BY e.event_id, e.start_date, e.end_date, category_name
            ORDER BY e.start_date;
        """

        result = await connection.fetch(base_query, *params)

        if not result:
            return {"error": "No data found for the specified range."}, 404

        # Prepare data for the response
        event_sales_data = [
            {
                "event_name": row["event_name"],
                "category_name": row["category_name"],
                "friendly_name": row["category_name"] + " at " + row["event_name"],
                "start_date": row["start_date"],
                "end_date": row["end_date"],
                "duration": int(row["duration"]),
                "average_sales_per_day": float(row["average_sales_per_day"]),
                "average_books_sold_per_day": int(row["average_books_sold_per_day"]),
                "total_sales": float(row["total_sales"]),
                "total_quantity_sold": int(row["total_quantity_sold"]),
                "unique_books_sold": int(row["unique_books_sold"])
            }
            for row in result
        ]

        return {"data": event_sales_data}

    except Exception as e:
        logging.error(f"Error: {e}")
        return {"error": f"An error occurred: {e}"}, 500
    
#3. API endpoint to export sales per event charts
@app.get('/api/sales/export-event-sales')
async def export_event_sales_plot():
    try:
        logging.debug("Establishing database connection...")
        connection = await asyncpg.connect(dsn=DB_URL)

        # Validate and parse query parameters
        start_date = request.args.get("startDate")
        end_date = request.args.get("endDate")
        category = request.args.get("category", None, type=int)

        # Ensure required parameters are present
        if not start_date or not end_date:
            return {"error": "startDate and endDate are required."}, 400

        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            if start_date > end_date:
                return {"error": "startDate must be before endDate."}, 400
        except ValueError:
            return {"error": "Invalid date format. Use YYYY-MM-DD."}, 400

        # Build and execute query to fetch event data and sales
        base_query = """
            SELECT 
                e.event_name,
                cat.category_name as category_name,
                e.start_date, 
                e.end_date, 
                CAST(e.end_date - e.start_date AS INTEGER) + 1 AS duration,
                SUM(s.quantity) AS total_quantity_sold,
                SUM(s.total_price) AS total_sales,
                COUNT(DISTINCT s.book_id) AS unique_books_sold,

                CASE 
                    WHEN (e.end_date - e.start_date) > 0 
                    THEN SUM(s.total_price) / (e.end_date - e.start_date) 
                    ELSE SUM(s.total_price)
                END AS average_sales_per_day,
                
                CASE 
                    WHEN (e.end_date - e.start_date) > 0 
                    THEN SUM(s.quantity) / (e.end_date - e.start_date) 
                    ELSE SUM(s.quantity)
                END AS average_books_sold_per_day
            FROM events e
            LEFT JOIN sales s ON s.event_id = e.event_id
            INNER JOIN books b ON b.book_id = s.book_id
            INNER JOIN subcategories sub ON b.subcategory_id = sub.subcategory_id
            LEFT JOIN categories cat ON sub.category_id = cat.category_id
            WHERE e.start_date BETWEEN $1 AND $2
        """

        # Add category filter dynamically
        params = [start_date, end_date]
        if category and category != 0:
            base_query += " AND cat.category_id = $3"
            params.append(category)

        # Group and order results
        base_query += """
            GROUP BY e.event_id, e.start_date, e.end_date, category_name
            ORDER BY e.start_date;
        """

        result = await connection.fetch(base_query, *params)

        if not result:
            return {"error": "No data found for the specified range."}, 404

        # Prepare data for the response
        event_sales_data = [
            {
                "event_name": row["event_name"],
                "category_name": row["category_name"],
                "friendly_name": row["category_name"] + " at " + row["event_name"],
                "start_date": row["start_date"],
                "end_date": row["end_date"],
                "duration": int(row["duration"]),
                "average_sales_per_day": float(row["average_sales_per_day"]),
                "average_books_sold_per_day": int(row["average_books_sold_per_day"]),
                "total_sales": float(row["total_sales"]),
                "total_quantity_sold": int(row["total_quantity_sold"]),
                "unique_books_sold": int(row["unique_books_sold"])
            }
            for row in result
        ]
        
        excel_output = create_separate_charts_with_duration(event_sales_data, "event_sales_data.xlsx")

        # Return the Excel file as a download
        return send_file(
            excel_output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="event_sales_data.xlsx"
        )

    except Exception as e:
        logging.error(f"Error: {e}")
        return {"error": f"An error occurred: {e}"}, 500

#3. Create bar chart + line chart for sales per event analysis
def create_separate_charts_with_duration(data, output_file):
    # Create a new workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Event Sales"

    # Add headers including Duration
    headers = [
        "Event name","Category name", "Friendly name", "Total sales", "Total quantity sold", 
        "Average sales per day", "Average books sold per day", "Unique books sold", "Duration (days)"
    ]
    ws.append(headers)

    # Add data to the sheet, including Duration
    for entry in data:
        ws.append([
            entry["event_name"],
            entry["category_name"],
            entry["friendly_name"],
            entry["total_sales"],
            entry["total_quantity_sold"],
            entry["average_sales_per_day"],
            entry["average_books_sold_per_day"],
            entry["unique_books_sold"],
            entry["duration"]
        ])

    # Chart 1: Total Sales and Average Sales Per Day
    # Create a Bar Chart (for Total Sales)
    bar_chart1 = BarChart()
    bar_chart1.type = "col"  # Clustered column chart
    bar_chart1.title = "Total sales and average sales per day"
    bar_chart1.x_axis.title = "Category sales at events"
    bar_chart1.y_axis.title = "Values"
    bar_chart1.width = 30
    bar_chart1.height = 15
    bar_chart1.gapWidth = 500
    bar_chart1.style = 10
    bar_chart1.x_axis.majorGridlines = None  # Remove gridlines

    # Define data and categories for the Bar Chart
    bar_data_ref1 = Reference(ws, min_col=4, max_col=4, min_row=1, max_row=len(data) + 1)
    categories_ref1 = Reference(ws, min_col=3, min_row=2, max_row=len(data) + 1)
    
    bar_chart1.add_data(bar_data_ref1, titles_from_data=True)
    bar_chart1.set_categories(categories_ref1)

    # Create a Line Chart (for Average Sales Per Day)
    line_chart1 = LineChart()
    line_data_ref1 = Reference(ws, min_col=6, max_col=6, min_row=1, max_row=len(data) + 1)
    line_chart1.add_data(line_data_ref1, titles_from_data=True)
    line_chart1.set_categories(categories_ref1)
    line_chart1.y_axis.axId = 200  # Assign a new Y-axis for the line chart
    line_chart1.x_axis = bar_chart1.x_axis  # Share the same X-axis with the bar chart
    line_chart1.title = None  # No separate title for the line chart
    line_chart1.width = 30
    line_chart1.height = 15
    line_chart1.style = 10
    # Combine the charts
    bar_chart1.y_axis.crosses = "autoZero"  # Keep bar chart Y-axis on the left
    bar_chart1 += line_chart1  # Add the line chart to the bar chart

    # Add the first chart to the worksheet
    ws.add_chart(bar_chart1, "K3")

    # Chart 2: Total Quantity Sold and Average Books Sold Per Day
    # Create a Bar Chart (for Total Quantity Sold)
    bar_chart2 = BarChart()
    bar_chart2.type = "col"  # Clustered column chart
    bar_chart2.title = "Total quantity sold and average books sold per day"
    bar_chart2.x_axis.title = "Category sales at events"
    bar_chart2.y_axis.title = "Values"
    bar_chart2.width = 30
    bar_chart2.height = 15
    bar_chart2.gapWidth = 500
    bar_chart2.style = 10
    
    # Define data and categories for the Bar Chart
    bar_data_ref2 = Reference(ws, min_col=5, max_col=5, min_row=1, max_row=len(data) + 1)
    categories_ref2 = Reference(ws, min_col=3, min_row=2, max_row=len(data) + 1)

    bar_chart2.add_data(bar_data_ref2, titles_from_data=True)
    bar_chart2.set_categories(categories_ref2)

    # Create a Line Chart (for Average Books Sold Per Day)
    line_chart2 = LineChart()
    line_data_ref2 = Reference(ws, min_col=7, max_col=7, min_row=1, max_row=len(data) + 1)
    line_chart2.add_data(line_data_ref2, titles_from_data=True)
    line_chart1.set_categories(categories_ref2)
    line_chart2.y_axis.axId = 300  # Assign a new Y-axis for the line chart
    line_chart2.x_axis = bar_chart2.x_axis  # Share the same X-axis with the bar chart
    line_chart2.title = None  # No separate title for the line chart
    line_chart2.width = 30
    line_chart2.height = 15
    line_chart2.style = 10
    # Combine the charts
    bar_chart2.y_axis.crosses = "autoZero"  # Keep bar chart Y-axis on the left
    bar_chart2 += line_chart2  # Add the line chart to the bar chart

    # Add the second chart to the worksheet
    ws.add_chart(bar_chart2, "K45")

    # Save the workbook
    wb.save(output_file)
    print(f"Excel file '{output_file}' with separate charts and duration created successfully!")
    return output_file



# API endpoint to fetch sales data grouped per city
@app.get("/api/sales/cities")
async def fetch_sales_by_city():
    try:
        logging.debug("Establishing database connection...")
        connection = await asyncpg.connect(dsn=DB_URL)

        # Parse query parameters
        start_date = request.args.get("startDate")
        end_date = request.args.get("endDate")
        if not start_date or not end_date:
            return {"error": "startDate and endDate are required."}, 400

        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            if start_date > end_date:
                return {"error": "startDate must be before endDate."}, 400
        except ValueError:
            return {"error": "Invalid date format. Use YYYY-MM-DD."}, 400

        # Query to fetch sales data grouped by city, gender, and age group
        query = """
            SELECT 
                c.city_name,
                c.latitude,
                c.longitude,
                SUM(s.total_price) AS total_sales,
                COUNT(s.sale_id) AS transaction_count,
                AVG(s.total_price) AS average_sale,
                MIN(s.total_price) AS min_sale,
                MAX(s.total_price) AS max_sale,
                COALESCE(gender, 'Unknown') AS gender,
                COALESCE(ag.age_group_name, 'Unknown') AS age_group,
                COALESCE(ag.description, 'Unknown') AS age_group_description,
                COUNT(*) AS group_count
            FROM sales s
            LEFT JOIN clients cl ON s.client_id = cl.client_id
            LEFT JOIN age_groups ag ON cl.age_group_id = ag.age_group_id
            LEFT JOIN cities c ON s.city_id = c.city_id
            WHERE s.sale_date BETWEEN $1 AND $2
            GROUP BY c.city_id, c.city_name, c.latitude, c.longitude, gender, ag.age_group_name, ag.description
            ORDER BY c.city_name, gender, age_group;
        """

        # Fetch data
        sales_data = await connection.fetch(query, start_date, end_date)

        # Aggregate the data by city
        city_data = {}
        for row in sales_data:
            city_name = row["city_name"]
            if city_name not in city_data:
                city_data[city_name] = {
                    "latitude": row["latitude"],
                    "longitude": row["longitude"],
                    "total_sales": 0,
                    "transaction_count": 0,
                    "average_sale": 0,
                    "min_sale": None,
                    "max_sale": None,
                    "gender_breakdown": {},
                    "age_group_distribution": {}
                }
            city = city_data[city_name]

            # Update overall city statistics
            city["total_sales"] += row["total_sales"]
            city["transaction_count"] += row["transaction_count"]
            city["average_sale"] = city["total_sales"] / city["transaction_count"]
            city["min_sale"] = min(filter(None, [city["min_sale"], row["min_sale"]]))
            city["max_sale"] = max(filter(None, [city["max_sale"], row["max_sale"]]))

            # Update gender breakdown
            gender = row["gender"]
            city["gender_breakdown"][gender] = city["gender_breakdown"].get(gender, 0) + row["group_count"]

            # Update age group distribution
            age_group = row["age_group"]
            age_group_description = row["age_group_description"]
            key = f"{age_group} ({age_group_description})"
            city["age_group_distribution"][key] = city["age_group_distribution"].get(key, 0) + row["group_count"]

        # Normalize percentages for gender and age group
        for city in city_data.values():
            total_count = city["transaction_count"]
            city["gender_breakdown"] = {k: round((v / total_count) * 100, 2) for k, v in city["gender_breakdown"].items()}
            city["age_group_distribution"] = {k: round((v / total_count) * 100, 2) for k, v in city["age_group_distribution"].items()}

        return jsonify(city_data)

    except Exception as e:
        logging.error(f"Error: {e}")
        return {"error": f"An error occurred while processing the request: {e}"}, 500

    finally:
        await connection.close()
  
if __name__ == "__main__":
    app.run(debug=True)
